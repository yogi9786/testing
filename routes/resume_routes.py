import base64
import datetime
import os
from fastapi import APIRouter, HTTPException, File, UploadFile, Form, Path
from fastapi.responses import FileResponse
from database import resume_collection
from models import Resume
from bson import ObjectId
from services.email_service import send_email
import pandas as pd
import openpyxl
from openpyxl.styles import Font

router = APIRouter()



@router.post("/upload/")
async def upload_resume(
    name: str = Form(...),
    phone: str = Form(...),
    email: str = Form(...),
    role: str = Form(...),
    resume: UploadFile = File(...),
):
    try:
        binary_data = await resume.read()
        encoded_resume = base64.b64encode(binary_data).decode("utf-8")

        resume_data = {
            "name": name,
            "phone": phone,
            "email": email,
            "role": role,
            "applied_at": datetime.datetime.now(datetime.timezone.utc),
            "resume": encoded_resume,
        }

        result = resume_collection.insert_one(resume_data)
        resume_id = str(result.inserted_id)

        resume_url = f"http://localhost:8000/download/{resume_id}"

        user_email_content = f"""
        <p>Hi {name},</p>
        <p>Thank you for applying for the job "<strong>{role}</strong>". We have received your application and will review it shortly.</p>
        <br>
        <p>Best regards,</p>
        <p><strong>Xtransmatrix Consulting Services Pvt Ltd</strong></p>
        """
        send_email(email, f"Thank you for Applying - {role} at xTransMatrix", user_email_content)

        return {"message": "Resume uploaded successfully", "id": resume_id}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@router.get("/resumes/", response_model=list[Resume])
def get_resumes():
    resumes = list(resume_collection.find({}, {"_id": 1, "name": 1, "phone": 1, "email": 1, "role": 1, "applied_at": 1}))

    return [
        Resume(
            id=str(resume["_id"]),
            name=resume.get("name", "N/A"),
            phone=resume.get("phone", "N/A"),
            email=resume.get("email", "N/A"),
            role=resume.get("role", "N/A"),
            applied_at=resume.get("applied_at", datetime.datetime.utcnow()),
        )
        for resume in resumes
    ]

@router.get("/download/{resume_id}")
def download_resume(resume_id: str):
    try:
        resume = resume_collection.find_one({"_id": ObjectId(resume_id)})
        if not resume:
            raise HTTPException(status_code=404, detail="Resume not found")

        binary_data = base64.b64decode(resume["resume"])
        file_path = f"temp_resume_{resume_id}.pdf"
        with open(file_path, "wb") as f:
            f.write(binary_data)

        return FileResponse(file_path, media_type="application/pdf", filename="resume.pdf")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/view/{resume_id}")
def view_resume(resume_id: str):
    try:
        resume = resume_collection.find_one({"_id": ObjectId(resume_id)}, {"resume": 1})
        if not resume or "resume" not in resume:
            raise HTTPException(status_code=404, detail="Resume not found")

        binary_data = base64.b64decode(resume["resume"])
        file_path = f"temp_resume_{resume_id}.pdf"

        with open(file_path, "wb") as f:
            f.write(binary_data)

        return FileResponse(file_path, media_type="application/pdf", headers={"Content-Disposition": "inline"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/delete/{resume_id}")
def delete_resume(resume_id: str):
    try:
        result = resume_collection.delete_one({"_id": ObjectId(resume_id)})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Resume not found")

        return {"message": "Resume deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/career/excel")
def export_users_to_excel():
    try:
        users = list(resume_collection.find({}, {"_id": 1, "name": 1, "phone": 1, "email": 1, "role": 1, "applied_at": 1}))

        if not users:
            raise HTTPException(status_code=404, detail="No data found to export.")

        base_url = "https://xtmx-career-backend-3.onrender.com"

        for user in users:
            user["id"] = str(user["_id"])
            del user["_id"]
            user["Download Link"] = f"{base_url}/download/{user['id']}"
            user["View Link"] = f"{base_url}/resume/{user['id']}"

        df = pd.DataFrame(users)

        file_path = "user_data.xlsx"
        df.to_excel(file_path, index=False, engine="openpyxl")

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        download_col = df.columns.get_loc("Download Link") + 1
        view_col = df.columns.get_loc("View Link") + 1

        for row in range(2, len(users) + 2):
            download_cell = ws.cell(row=row, column=download_col)
            view_cell = ws.cell(row=row, column=view_col)

            download_cell.hyperlink = download_cell.value
            download_cell.font = Font(color="0000FF", underline="single")

            view_cell.hyperlink = view_cell.value
            view_cell.font = Font(color="0000FF", underline="single")

        wb.save(file_path)

        return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="user_data.xlsx")

    except Exception as e:
        print("Error exporting data:", e)
        raise HTTPException(status_code=500, detail="Failed to generate Excel file.")
