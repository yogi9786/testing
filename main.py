import datetime
import io
import os
from fastapi import FastAPI, HTTPException, File, UploadFile, Form, Response
from pydantic import BaseModel, EmailStr
from pymongo.mongo_client import MongoClient
from pymongo.server_api import ServerApi
from dotenv import load_dotenv
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.base import BaseHTTPMiddleware
import traceback
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
import base64
from bson import ObjectId
from typing import List, Optional
from fastapi.params import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

load_dotenv()

MONGO_URI = os.getenv("MONGO_URI")
client = MongoClient(MONGO_URI, server_api=ServerApi('1'))

# Defining different databases and collections
contact_db = client["contact_database"]
resume_db = client["resume_database"]

contact_collection = contact_db["contact_forms"]
resume_collection = resume_db["resumes"]

# Updating missing fields in the database
resume_collection.update_many({"phone": None}, {"$set": {"phone": ""}})
resume_collection.update_many({"resume": None}, {"$set": {"resume": ""}})

if os.getenv("SENDGRID_API_KEY"):
    print("SENDGRID_API_KEY loaded successfully!")
else:
    print("SENDGRID_API_KEY is missing!")
SENDGRID_API_KEY = os.getenv("SENDGRID_API_KEY")
FROM_EMAIL = os.getenv("FROM_EMAIL")
ADMIN_EMAIL = os.getenv("ADMIN_EMAIL")

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

class IgnoreFaviconMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        if request.url.path == "/favicon.ico":
            return Response(status_code=204)  
        return await call_next(request)

app.add_middleware(IgnoreFaviconMiddleware)

def get_ist_time():
    """Returns the current date in IST (Indian Standard Time, UTC+5:30)."""
    return (datetime.datetime.utcnow() + datetime.timedelta(hours=5, minutes=30)).date()
    


class ContactForm(BaseModel):
    name: str
    email: EmailStr
    message: str

class Resume(BaseModel):
    id: Optional[str] = None  # Make id optional
    name: str
    phone: Optional[str] = None
    email: Optional[str] = None
    role: str
    applied_at: datetime.datetime
    resume: Optional[str] = None

    
def get_ist_time():
    """Returns the current date in IST (Indian Standard Time, UTC+5:30)."""
    return (datetime.datetime.utcnow() + datetime.timedelta(hours=5, minutes=30)).date()
    
def send_email(to_email: str, name: str, message: str):
    """Function to send email using SendGrid."""
    try:
        email_content = f"""
        <html>
            <body>
                <h2>Hello {name},</h2>
                <p>Thank you for reaching out to us!</p>
                <p>Your message: {message}</p>
                <p>We will get back to you soon.</p>
                <br>
                <p>Best Regards,<br>XTRANSMATRIX CONSULTING SERVICES PVT LTD</p>
            </body>
        </html>
        """

        mail = Mail(
            from_email=FROM_EMAIL,
            to_emails=to_email,
            subject="Thank you for contacting us!",
            html_content=email_content
        )

        sg = SendGridAPIClient(SENDGRID_API_KEY)
        response = sg.send(mail)
        return response.status_code
    except Exception as e:
        print("SendGrid Error:", traceback.format_exc())


def send_email(to_email: str, subject: str, content: str):
    try:
        message = Mail(
            from_email="yogesh.v@xtransmatrix.com", 
            to_emails=to_email,
            subject=subject,
            html_content=content,
        )
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        sg.send(message)
    except Exception as e:
        print(f"Error sending email: {e}")

@app.get("/")
def read_root():
    return {"message": "FastAPI backend is running successfully with MongoDB Atlas!"}

# Contact Form Submission API
@app.post("/submit")
def submit_contact_form(form: ContactForm):
    try:
        form_data = form.dict()
        result = contact_collection.insert_one(form_data)

        email_content = f"""
        <html>
            <body>
                <h2>Hello {form.name},</h2>
                <p>Thank you for reaching out to us!</p>
                <p>Your message: {form.message}</p>
                <p>We will get back to you soon.</p>
                <br>
                <p>Best Regards,<br>XTRANSMATRIX CONSULTING SERVICES PVT LTD</p>
            </body>
        </html>
        """

        mail = Mail(
            from_email=FROM_EMAIL,
            to_emails=form.email,
            subject="Thank you for contacting us!",
            html_content=email_content
        )

        sg = SendGridAPIClient(SENDGRID_API_KEY)
        response = sg.send(mail)
        return {"message": "Contact form submitted successfully", "id": str(result.inserted_id)}
    except Exception as e:
        print("Error:", traceback.format_exc())  
        raise HTTPException(status_code=500, detail="Internal Server Error. Check logs for details.")

# Get all contact form submissions
@app.get("/submissions", response_model=List[dict])
def get_contact_submissions():
    try:
        submissions = list(contact_collection.find({}, {"_id": 1, "name": 1, "email": 1, "message": 1}))

        return [
            {
                "id": str(sub["_id"]),
                "name": sub.get("name", "N/A"),
                "email": sub.get("email", "N/A"),
                "message": sub.get("message", "N/A"),
            }
            for sub in submissions
        ]
    except Exception as e:
        raise HTTPException(status_code=500, detail="Internal Server Error. Check logs for details.")
    
@app.get("/contacts-excel", summary="Export Contacts Data to Excel", tags=["Export"])
def export_contacts_to_excel():
    """
    Fetches all contact data from `contact_collection` and exports it as an Excel file.
    """
    try:
        contacts = list(contact_collection.find({}, {"_id": 0}))  

        if not contacts:
            raise HTTPException(status_code=404, detail="No contact data found to export.")

        df = pd.DataFrame(contacts)
        file_path = "contact_data.xlsx"
        df.to_excel(file_path, index=False, engine="openpyxl")

        return FileResponse(
            file_path,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename="contact_data.xlsx"
        )

    except Exception as e:
        print("Error exporting contact data:", e)
        raise HTTPException(status_code=500, detail="Failed to generate contact Excel file.")

# Delete a contact submission
@app.delete("/delete/{submission_id}")
def delete_contact_submission(submission_id: str = Path(..., title="Submission ID")):
    try:
        if not ObjectId.is_valid(submission_id):
            raise HTTPException(status_code=400, detail="Invalid submission ID format")
        
        result = contact_collection.delete_one({"_id": ObjectId(submission_id)})

        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Submission not found")

        return {"message": "Submission deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to delete submission: {str(e)}")


@app.post("/upload/", response_model=dict)
async def upload_resume(
    name: str = Form(...),
    phone: str = Form(...),
    email: EmailStr = Form(...),
    role: str = Form(...),
    resume: UploadFile = File(...),
):
    try:
        # Encode Resume File
        binary_data = await resume.read()
        encoded_resume = base64.b64encode(binary_data).decode("utf-8")

        # Store in MongoDB
        resume_data = {
            "name": name,
            "phone": phone,
            "email": email,
            "role": role,
            "applied_at": datetime.datetime.now(datetime.timezone.utc),
            "resume": encoded_resume,
        }
        result = resume_collection.insert_one(resume_data)
        resume_id = str(result.inserted_id)  # Convert ObjectId to string

        # Generate Resume Download URL
        resume_url = f"http://localhost:8000/download/{resume_id}"  # Update with your hosted URL

        # Send Email to User
        user_email_content = f"""
        <p>Hi {name},</p>
        <p>Thank you for applying for the job "<strong>{role}</strong>". We have received your application and will review it shortly.</p>
        <br>
        <p>Best regards,</p>
        <p><strong>Xtransmatrix Consulting Services Pvt Ltd</strong></p>
        """
        send_email(email, f"Thank you for Applying - {role} at xTransMatrix", user_email_content)

        # Send Email to Admin
        admin_email_content = f"""
        <p><strong>New Job Application Received</strong></p>
        <p><strong>Name:</strong> {name}</p>
        <p><strong>Phone:</strong> {phone}</p>
        <p><strong>Email:</strong> {email}</p>
        <p><strong>Date:</strong> {datetime.datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')}</p>
        <p><strong>Role:</strong> {role}</p>
        <p><strong>Resume:</strong> <a href="{resume_url}" target="_blank" style="text-decoration: none; color: #4CAF50;">View Resume</a></p>
        """
        send_email(ADMIN_EMAIL, f"New Application for {role}", admin_email_content)

        return {"message": "Resume uploaded successfully", "id": resume_id}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/resumes/", response_model=List[Resume])
def get_resumes():
    resumes = list(resume_collection.find({}, {"_id": 1, "name": 1, "phone": 1, "email": 1, "role": 1, "applied_at": 1}))

    return [
        Resume(
            id=str(resume["_id"]),  # Convert ObjectId to string
            name=resume.get("name", "N/A"),
            phone=resume.get("phone", "N/A"),
            email=resume.get("email", "N/A"),
            role=resume.get("role", "N/A"),
            applied_at=resume.get("applied_at", datetime.datetime.utcnow()),
        )
        for resume in resumes
    ]

# Get a specific resume
@app.get("/resume/{resume_id}", response_model=Resume)
def get_resume(resume_id: str):
    try:
        resume = resume_collection.find_one({"_id": ObjectId(resume_id)}, {"_id": 1, "name": 1, "phone": 1, "email": 1, "role": 1, "applied_at": 1})
        if not resume:
            raise HTTPException(status_code=404, detail="Resume not found")
        
        return Resume(
            id=str(resume["_id"]),  # Ensure `id` is returned
            name=resume["name"],
            phone=resume.get("phone", None),
            email=resume["email"],
            role=resume["role"],
            applied_at=resume["applied_at"],
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

# Download Resume
@app.get("/download/{resume_id}")
def download_resume(resume_id: str):
    try:
        resume = resume_collection.find_one({"_id": ObjectId(resume_id)})
        if not resume:
            raise HTTPException(status_code=404, detail="Resume not found")

        binary_data = base64.b64decode(resume["resume"])
        file_path = f"temp_resume_{resume_id}.pdf"
        with open(file_path, "wb") as f:
            f.write(binary_data)

        return FileResponse(file_path, media_type="application/pdf", filename="resume.pdf")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/view/{resume_id}")
def view_resume(resume_id: str):
    try:
        resume = resume_collection.find_one({"_id": ObjectId(resume_id)})
        if not resume:
            raise HTTPException(status_code=404, detail="Resume not found")

        binary_data = base64.b64decode(resume["resume"])
        file_path = f"temp_resume_{resume_id}.pdf"
        with open(file_path, "wb") as f:
            f.write(binary_data)

        # Return file with inline disposition (opens in browser)
        return FileResponse(file_path, media_type="application/pdf", headers={"Content-Disposition": "inline; filename=resume.pdf"})
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@app.delete("/delete/{resume_id}")
def delete_resume(resume_id: str):
    try:
        result = resume_collection.delete_one({"_id": ObjectId(resume_id)})
        
        if result.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Resume not found")
        
        return {"message": "Resume deleted successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/career/excel")
def export_users_to_excel():
    try:
        users = list(resume_collection.find({}, {"_id": 1, "name": 1, "phone": 1, "email": 1, "role": 1, "applied_at": 1}))  

        if not users:
            raise HTTPException(status_code=404, detail="No data found to export.")

        base_url = "https://xtmx-career-backend-3.onrender.com"

        # Prepare data for DataFrame
        for user in users:
            user["id"] = str(user["_id"])
            del user["_id"]  # Remove ObjectId field
            user["Download Link"] = f"{base_url}/download/{user['id']}"
            user["View Link"] = f"{base_url}/resume/{user['id']}"

        df = pd.DataFrame(users)

        file_path = "user_data.xlsx"
        df.to_excel(file_path, index=False, engine="openpyxl")  

        # Open the Excel file and format the hyperlinks
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Find column indices for Download and View Links
        download_col = df.columns.get_loc("Download Link") + 1
        view_col = df.columns.get_loc("View Link") + 1

        # Apply hyperlink format
        for row in range(2, len(users) + 2):  # Start from row 2 (Excel headers in row 1)
            download_cell = ws.cell(row=row, column=download_col)
            view_cell = ws.cell(row=row, column=view_col)

            download_cell.hyperlink = download_cell.value
            download_cell.font = Font(color="0000FF", underline="single")

            view_cell.hyperlink = view_cell.value
            view_cell.font = Font(color="0000FF", underline="single")

        wb.save(file_path)  # Save the updated file

        return FileResponse(file_path, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename="user_data.xlsx")

    except Exception as e:
        print("Error exporting data:", e)
        raise HTTPException(status_code=500, detail="Failed to generate Excel file.")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
